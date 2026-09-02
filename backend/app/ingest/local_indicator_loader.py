"""Load LCFF Local Indicator files, from spreadsheets or text.

Only ~2,300 rows per file, so this does not need the ``COPY`` machinery the
state indicators use; it stages through a temporary table for the same reason
they do -- so a reload replaces rather than duplicates -- and otherwise keeps
the code plain.

Spreadsheets are the preferred source.  The text exports are byte-identical in
content but replace the newlines inside narrative fields with spaces, and
Priority 3's narratives run to thousands of words of deliberately structured
prose.  See :doc:`/data/local-indicators`.
"""

from __future__ import annotations

import logging
import time
import uuid
from collections.abc import Iterable, Iterator, Sequence
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from psycopg.types.json import Json
from sqlalchemy import Engine, text
from sqlmodel import Session, col, select

from app.ingest.local_indicator_parser import (
    LocalIndicatorLayoutError,
    LocalIndicatorRecord,
    LocalIndicatorRowParser,
    detect_delimiter,
    iter_rows,
    parse_filename,
)
from app.ingest.local_indicator_reference import (
    PRIORITY_NUMBERS,
    seed_local_indicator_reference,
)
from app.ingest.parser import ParseError
from app.ingest.sources import (
    DASHBOARD_BASE_URL,
    HttpSource,
    ResearchFileSource,
    SourceObject,
    source_from_uri,
)
from app.model.ingest import IngestFile, IngestRun, IngestStatus

logger = logging.getLogger(__name__)

_COLUMNS = (
    "cds_code",
    "reporting_year",
    "priority_number",
    "lea_name",
    "performance",
    "meeting_date",
    "additional_info",
    "responses",
)

# The state has published local indicators every year since 2018 except 2020.
FIRST_YEAR = 2018
SUSPENDED_YEARS = frozenset({2020})


def local_indicator_names(
    years: Sequence[int] | None = None,
    priorities: Sequence[int] | None = None,
    *,
    suffix: str = ".xlsx",
) -> list[str]:
    """Every published Local Indicator file name, e.g. ``Pr32025.xlsx``."""
    span = years or [
        year
        for year in range(FIRST_YEAR, datetime.now(tz=UTC).year + 1)
        if year not in SUSPENDED_YEARS
    ]
    wanted = priorities or PRIORITY_NUMBERS
    return [f"Pr{p}{year}{suffix}" for year in span for p in wanted]


def read_spreadsheet_rows(path: Path) -> Iterator[list[Any]]:
    """Yield every row of a spreadsheet, header first."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        for row in sheet.iter_rows(values_only=True):
            # ``read_only`` sheets report a dimension that runs to the bottom
            # of the sheet, so blank trailing rows have to be skipped here.
            if row is None or all(cell is None for cell in row):
                continue
            yield list(row)
    finally:
        workbook.close()


def _row_values(record: LocalIndicatorRecord) -> tuple[Any, ...]:
    return (
        record.cds_code,
        record.reporting_year,
        record.priority_number,
        record.lea_name,
        record.performance,
        record.meeting_date,
        record.additional_info,
        Json(record.responses),
    )


@dataclass(slots=True)
class LoadCounts:
    """How much one file contributed."""

    rows: int = 0
    keys: set[tuple[int, int]] = field(default_factory=set)


class LocalIndicatorLoader:
    """Loads one Local Indicator file's parsed rows."""

    def __init__(self, engine: Engine) -> None:
        self.engine = engine

    def load(self, records: Iterable[LocalIndicatorRecord]) -> LoadCounts:
        counts = LoadCounts()
        # A file is one LEA per row, so the whole thing fits in memory
        # comfortably and can be de-duplicated before it reaches the database.
        by_key: dict[tuple[str, int, int], LocalIndicatorRecord] = {}
        for record in records:
            counts.keys.add((record.reporting_year, record.priority_number))
            by_key[(record.cds_code, record.reporting_year, record.priority_number)] = (
                record
            )

        if not by_key:
            return counts

        columns = ", ".join(_COLUMNS)
        placeholders = ", ".join("%s" for _ in _COLUMNS)
        with self.engine.begin() as connection:
            driver = connection.connection.driver_connection
            if driver is None:
                raise RuntimeError("no DBAPI connection behind the engine")
            with driver.cursor() as cursor:
                years = sorted({year for year, _ in counts.keys})
                priorities = sorted({p for _, p in counts.keys})
                cursor.execute(
                    "DELETE FROM local_indicator_results "
                    "WHERE reporting_year = ANY(%s) AND priority_number = ANY(%s)",
                    (years, priorities),
                )
                # Rows whose LEA is not a known entity are dropped rather than
                # failing the load; two of the ~2,300 are not in the entity
                # dimension, and losing the file over them helps nobody.
                cursor.execute(
                    "SELECT cds_code FROM entities WHERE cds_code = ANY(%s)",
                    ([key[0] for key in by_key],),
                )
                known = {row[0] for row in cursor.fetchall()}
                rows = [
                    _row_values(record)
                    for key, record in by_key.items()
                    if key[0] in known
                ]
                skipped = len(by_key) - len(rows)
                if skipped:
                    logger.info("skipping %s rows with no matching entity", skipped)
                cursor.executemany(
                    f"INSERT INTO local_indicator_results ({columns}) "
                    f"VALUES ({placeholders})",
                    rows,
                )
                counts.rows = len(rows)
        return counts


def analyze(engine: Engine) -> None:
    """Refresh planner statistics after a load."""
    with engine.connect().execution_options(isolation_level="AUTOCOMMIT") as connection:
        connection.execute(text("ANALYZE local_indicator_results"))


@dataclass(slots=True)
class FileOutcome:
    """What happened to one file during a run."""

    name: str
    status: IngestStatus
    priority: int | None = None
    reporting_year: int | None = None
    rows: int = 0
    duration_seconds: float = 0.0
    error: str | None = None


@dataclass(slots=True)
class RunOutcome:
    """The result of a whole run."""

    run_id: str
    source_uri: str
    status: IngestStatus
    files: list[FileOutcome] = field(default_factory=list)

    @property
    def rows(self) -> int:
        return sum(outcome.rows for outcome in self.files)


class LocalIndicatorImportRunner:
    """Runs Local Indicator imports and records what it did."""

    #: Spreadsheets preserve the paragraph breaks the text exports flatten.
    suffix = ".xlsx"

    def __init__(self, engine: Engine) -> None:
        self.engine = engine
        self.loader = LocalIndicatorLoader(engine)

    def run(
        self,
        source_uri: str = DASHBOARD_BASE_URL,
        *,
        force: bool = False,
        years: Sequence[int] | None = None,
        priorities: Sequence[int] | None = None,
    ) -> RunOutcome:
        """Import every changed Local Indicator file under ``source_uri``."""
        source = source_from_uri(source_uri)
        if isinstance(source, HttpSource):
            source.names = tuple(
                local_indicator_names(years, priorities, suffix=self.suffix)
            )

        with Session(self.engine) as session:
            seed_local_indicator_reference(session)
            session.commit()
            run = IngestRun(source_uri=source.uri, started_at=datetime.now(tz=UTC))
            session.add(run)
            session.commit()
            session.refresh(run)
            run_id = run.id

        outcome = RunOutcome(
            run_id=str(run_id), source_uri=source.uri, status=IngestStatus.RUNNING
        )
        try:
            for obj in source.list_objects():
                priority, year = parse_filename(obj.name)
                if priority is None:
                    continue
                if years and year is not None and year not in years:
                    continue
                if priorities and priority not in priorities:
                    continue
                outcome.files.append(
                    self._load_object(source, obj, run_id, force=force)
                )
            outcome.status = (
                IngestStatus.FAILED
                if any(f.status is IngestStatus.FAILED for f in outcome.files)
                else IngestStatus.SUCCEEDED
            )
        except Exception as error:  # noqa: BLE001 - recorded on the run row
            outcome.status = IngestStatus.FAILED
            self._finish_run(run_id, outcome, error=str(error))
            raise
        else:
            if outcome.rows:
                analyze(self.engine)
            self._finish_run(run_id, outcome)
        return outcome

    def _already_loaded(self, session: Session, obj: SourceObject) -> bool:
        previous = session.exec(
            select(IngestFile)
            .where(IngestFile.source_key == obj.key)
            .where(IngestFile.status == IngestStatus.SUCCEEDED)
            .order_by(col(IngestFile.loaded_at).desc())
            .limit(1)
        ).first()
        if previous is None:
            return False
        return (
            previous.etag == obj.etag
            and previous.size_bytes == obj.size_bytes
            and previous.result_rows > 0
        )

    def _records(
        self, source: ResearchFileSource, obj: SourceObject
    ) -> list[LocalIndicatorRecord]:
        priority, year = parse_filename(obj.name)
        suffix = Path(obj.name).suffix.lower()

        if suffix == ".xlsx":
            with source.open_file(obj) as path:
                rows = read_spreadsheet_rows(path)
                header = next(rows, None)
                if header is None:
                    raise LocalIndicatorLayoutError(f"{obj.name} is empty")
                parser = LocalIndicatorRowParser(
                    header, default_priority=priority, default_year=year
                )
                return [parser.parse(row) for row in rows]

        with source.open_text(obj) as lines:
            header_line = next(lines, "").rstrip("\r\n")
            if not header_line:
                raise LocalIndicatorLayoutError(f"{obj.name} is empty")
            delimiter = detect_delimiter(header_line)
            parser = LocalIndicatorRowParser(
                header_line.split(delimiter),
                default_priority=priority,
                default_year=year,
            )
            return [parser.parse(row) for row in iter_rows(lines, delimiter)]

    def _load_object(
        self,
        source: ResearchFileSource,
        obj: SourceObject,
        run_id: uuid.UUID,
        *,
        force: bool,
    ) -> FileOutcome:
        started = time.monotonic()
        priority, year = parse_filename(obj.name)

        with Session(self.engine) as session:
            if not force and self._already_loaded(session, obj):
                logger.info("skipping unchanged %s", obj.name)
                self._record_file(
                    session, run_id, obj, IngestStatus.SKIPPED, started=started
                )
                return FileOutcome(name=obj.name, status=IngestStatus.SKIPPED)

        try:
            counts = self.loader.load(self._records(source, obj))
        except (LocalIndicatorLayoutError, ParseError, ValueError) as error:
            logger.warning("skipping %s: %s", obj.name, error)
            with Session(self.engine) as session:
                self._record_file(
                    session,
                    run_id,
                    obj,
                    IngestStatus.FAILED,
                    started=started,
                    error=str(error),
                )
            return FileOutcome(
                name=obj.name, status=IngestStatus.FAILED, error=str(error)
            )

        duration = time.monotonic() - started
        with Session(self.engine) as session:
            self._record_file(
                session,
                run_id,
                obj,
                IngestStatus.SUCCEEDED,
                started=started,
                rows=counts.rows,
                priority=priority,
                year=year,
            )
        logger.info("loaded %s: %s rows", obj.name, f"{counts.rows:,}")
        return FileOutcome(
            name=obj.name,
            status=IngestStatus.SUCCEEDED,
            priority=priority,
            reporting_year=year,
            rows=counts.rows,
            duration_seconds=duration,
        )

    def _record_file(
        self,
        session: Session,
        run_id: uuid.UUID,
        obj: SourceObject,
        status: IngestStatus,
        *,
        started: float,
        rows: int = 0,
        priority: int | None = None,
        year: int | None = None,
        error: str | None = None,
    ) -> None:
        session.add(
            IngestFile(
                run_id=run_id,
                source_key=obj.key,
                etag=obj.etag,
                size_bytes=obj.size_bytes,
                last_modified=obj.last_modified,
                program="LOCAL",
                test_type=f"Pr{priority}" if priority else None,
                test_year=year,
                status=status,
                result_rows=rows,
                subscore_rows=0,
                duration_seconds=time.monotonic() - started,
                error=error,
                loaded_at=datetime.now(tz=UTC),
            )
        )
        session.commit()

    def _finish_run(
        self, run_id: uuid.UUID, outcome: RunOutcome, *, error: str | None = None
    ) -> None:
        with Session(self.engine) as session:
            run = session.get(IngestRun, run_id)
            if run is None:
                return
            run.status = outcome.status
            run.finished_at = datetime.now(tz=UTC)
            run.files_seen = len(outcome.files)
            run.files_loaded = sum(
                1 for f in outcome.files if f.status is IngestStatus.SUCCEEDED
            )
            run.files_skipped = sum(
                1 for f in outcome.files if f.status is IngestStatus.SKIPPED
            )
            run.result_rows = outcome.rows
            run.subscore_rows = 0
            run.error = error
            session.add(run)
            session.commit()
